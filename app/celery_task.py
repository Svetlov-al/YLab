import asyncio
import os
from datetime import timedelta

import httpx
import openpyxl
from celery import Celery

# Получаем текущую директорию скрипта
current_directory = os.path.dirname(__file__)

# Строим путь к файлу относительно текущей директории
file_path = os.path.join(current_directory, '..', 'admin', 'Menu.xlsx')

celery_app = Celery('ylab')

# Настройки брокера и другие
celery_app.conf.broker_url = 'amqp://guest:guest@rabbitmq:5672//'
celery_app.conf.result_backend = 'rpc://'
celery_app.conf.task_serializer = 'json'
celery_app.conf.result_serializer = 'json'
celery_app.conf.accept_content = ['json']
celery_app.conf.broker_connection_retry_on_startup = True

# Настройки Celery Beat
celery_app.conf.beat_schedule = {
    'database_synchronization': {
        'task': 'app.celery_task.database_synchronization',
        'schedule': timedelta(seconds=15),
    },
}


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    menus = []
    submenus = []
    dishes = []

    row = 1
    while row <= sheet.max_row:
        if sheet.cell(row=row, column=1).value and isinstance(sheet.cell(row=row, column=1).value, str):  # Новое меню
            menu = {
                'uuid': sheet.cell(row=row, column=1).value,
                'title': sheet.cell(row=row, column=2).value,
                'description': sheet.cell(row=row, column=3).value
            }
            menus.append(menu)
            row += 1

            while row <= sheet.max_row and not isinstance(sheet.cell(row=row, column=1).value, str):  # Подменю
                if sheet.cell(row=row, column=2).value and isinstance(sheet.cell(row=row, column=2).value, str):
                    submenu = {
                        'uuid': sheet.cell(row=row, column=2).value,
                        'title': sheet.cell(row=row, column=3).value,
                        'description': sheet.cell(row=row, column=4).value,
                        'menu_uuid': menu['uuid']
                    }
                    submenus.append(submenu)
                    row += 1

                    while row <= sheet.max_row and not isinstance(sheet.cell(row=row, column=1).value,
                                                                  str) and not isinstance(
                            sheet.cell(row=row, column=2).value, str):  # Блюдо
                        price_value = sheet.cell(row=row, column=6).value
                        if isinstance(price_value, str):
                            price_value = price_value.replace(',', '.')
                        dish = {
                            'uuid': sheet.cell(row=row, column=3).value,
                            'title': sheet.cell(row=row, column=4).value,
                            'description': sheet.cell(row=row, column=5).value,
                            'price': float(price_value or 0),
                            'submenu_uuid': submenu['uuid'],
                            'menu_uuid': menu['uuid']  # Добавляем информацию о меню
                        }
                        dishes.append(dish)
                        row += 1
                else:
                    row += 1
        else:
            row += 1

    return menus, submenus, dishes


@celery_app.task
def database_synchronization():
    # Используем созданный событийный цикл для выполнения асинхронной функции
    loop = asyncio.get_event_loop()
    menus_data, submenus_data, dishes_data = read_excel(file_path)
    loop.run_until_complete(async_database_operations(menus_data, submenus_data, dishes_data))


BASE_URL = 'http://api:8000/api/v1'
data = ([{'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40100',
          'title': 'Меню', 'description': 'Основное меню'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40102',
          'title': 'Алкогольное меню', 'description': 'Алкогольные напитки'}],
        [{'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40101',
          'title': 'Холодные закуски', 'description': 'К пиву',
          'menu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40100'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40103',
            'title': 'Рамен', 'description': 'Горячий рамен',
            'menu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40100'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40104',
          'title': 'Красные вина', 'description': 'Для романтичного вечера',
          'menu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40102'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40105',
          'title': 'Виски', 'description': 'Для интересных бесед',
          'menu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40102'}],
        [{'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40106',
          'title': 'Сельдь Бисмарк', 'description': 'Традиционное немецкое блюдо из маринованной сельди',
          'price': 182.99, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40101'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40107',
            'title': 'Мясная тарелка', 'description': 'Нарезка из ветчины, колбасных колечек,'
            ' нескольких сортов сыра и фруктов',
            'price': 215.36, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40101'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40108', 'title': 'Рыбная тарелка',
            'description': 'Нарезка из креветок, кальмаров, раковых шеек, гребешков,'
            ' лосося, скумбрии и красной икры', 'price': 265.57,
            'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40101'},
         {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40109',
            'title': 'Дайзу рамен', 'description': 'Рамен на курином бульоне с куриными подушками и яйцом аджитама,'
            ' яично-пшеничной лапшой, ростки зелени, грибами муэр и зеленым луком', 'price': 166.47, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40103'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40110', 'title': 'Унаги рамен', 'description': 'Рамен на нежном сливочном рыбном бульоне, с добавлением маринованного угря, грибов муэр, кунжута, зеленого лука', 'price': 168.25, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40103'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40111', 'title': 'Чиизу Рамен', 'description': 'Рамен на насыщенном сырном бульоне на основе кокосового молока, с добавлением куриной грудинки, яично - пшеничной лапши, мисо-матадоре, ростков зелени, листьев вакамэ', 'price': 132.88, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40103'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40112', 'title': 'Шемен де Пап ля Ноблесс', 'description': 'Вино красное — фруктовое, среднетелое, выдержанное в дубе', 'price': 2700.79, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40104'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40113', 'title': 'Рипароссо Монтепульчано', 'description': 'Вино красное, сухое', 'price': 3100.33, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40104'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40114', 'title': 'Кьянти, Серристори', 'description': 'Вино красное — элегантное, комплексное, не выдержанное в дубе', 'price': 1850.42, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40104'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40115', 'title': 'Джемисон', 'description': 'Классический купажированный виски, проходящий 4-хлетнюю выдержку в дубовых бочках', 'price': 420.78, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40105'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40116', 'title': 'Джек Дэниелс', 'description': 'Характерен мягкий вкус, сочетает в себе карамельно-ванильные и древесные нотки. Легкий привкус дыма.', 'price': 440.11, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40105'}, {'uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40117', 'title': 'Чивас Ригал', 'description': 'Это купаж высококачественных солодовых и зерновых виски, выдержанных как минимум в течение 12 лет, что придает напитку роскошные нотки меда, ванили и спелых яблок.', 'price': 520.08, 'submenu_uuid': '9bc1fc99-c53e-42ae-adec-6fe466b40105'}])


async def async_database_operations(menus_data, submenus_data, dishes_data):
    async with httpx.AsyncClient() as client:
        # Обработка меню
        for menu_data in menus_data:
            menu_id = menu_data['uuid']
            response = await client.get(f'{BASE_URL}/menus/{menu_id}')

            if response.status_code == 200:  # Если меню с таким UUID найдено
                # Удаляем uuid из данных, так как при обновлении мы передаем только title и description
                del menu_data['uuid']
                await client.patch(f'{BASE_URL}/menus/{menu_id}', json=menu_data)
            elif response.status_code == 404:  # Если меню с таким UUID не найдено
                # Переименовываем ключ "uuid" в "id" для создания нового меню
                menu_data['id'] = menu_data.pop('uuid')
                await client.post(f'{BASE_URL}/menus', json=menu_data)

        # Обработка подменю
        for submenu_data in submenus_data:
            # print(f"DATAAAAAA{submenu_data}")
            submenu_uuid = submenu_data['uuid']
            menu_uuid = submenu_data['menu_uuid']
            response = await client.get(f'{BASE_URL}/menus/{menu_uuid}/submenus/{submenu_uuid}')

            if response.status_code == 200:  # Если подменю с таким UUID найдено
                del submenu_data['uuid']
                del submenu_data['menu_uuid']
                await client.patch(f'{BASE_URL}/menus/{menu_uuid}/submenus/{submenu_uuid}', json=submenu_data)
            elif response.status_code == 404:  # Если подменю с таким UUID не найдено
                submenu_data['id'] = submenu_data.pop('uuid')
                del submenu_data['menu_uuid']
                await client.post(f'{BASE_URL}/menus/{menu_uuid}/submenus', json=submenu_data)

        # Обработка блюд
        for dish_data in dishes_data:
            dish_uuid = dish_data['uuid']
            menu_uuid = dish_data['menu_uuid']
            submenu_uuid = dish_data['submenu_uuid']
            response = await client.get(f'{BASE_URL}/menus/{menu_uuid}/submenus/{submenu_uuid}/dishes/{dish_uuid}')

            if response.status_code == 200:  # Если блюдо с таким UUID найдено
                del dish_data['uuid']
                del dish_data['menu_uuid']
                del dish_data['submenu_uuid']
                await client.patch(f'{BASE_URL}/menus/{menu_uuid}/submenus/{submenu_uuid}/dishes/{dish_uuid}', json=dish_data)
            elif response.status_code == 404:  # Если блюдо с таким UUID не найдено
                dish_data['id'] = dish_data.pop('uuid')
                del dish_data['menu_uuid']
                del dish_data['submenu_uuid']
                await client.post(f'{BASE_URL}/menus/{menu_uuid}/submenus/{submenu_uuid}/dishes', json=dish_data)

# async def async_database_operations(data_from_excel):
#     async with AsyncSessionLocal() as db:
#         menu_repo = MenuRepository(db)
#         submenu_repo = SubmenuRepository(db)
#         dish_repo = DishRepository(db)
#         existing_menus = await menu_repo.get_full_menus()
#
#         for menu_data in data_from_excel:
#             menu_title = menu_data["title"]
#             existing_menu = next((menu for menu in existing_menus if menu.title == menu_title), None)
#
#             if not existing_menu:
#                 # Создание нового меню
#                 menu_obj = schemas.MenuCreate(title=menu_data["title"], description=menu_data["description"])
#                 new_menu = await menu_repo.create_menu(menu_obj)
#             else:
#                 # Обновление существующего меню
#                 menu_obj = schemas.MenuCreate(title=menu_data["title"], description=menu_data["description"])
#                 new_menu = await menu_repo.update_menu(existing_menu.id, menu_obj)
#
#             # Обработка подменю
#             submenus = menu_data.get("submenus", [])
#             for submenu_data in submenus:
#                 submenu_title = submenu_data["title"]
#                 existing_submenu = None
#                 if existing_menu:
#                     existing_submenu = next(
#                         (submenu for submenu in existing_menu.submenus if submenu.title == submenu_title), None)
#
#                 if not existing_submenu:
#                     # Создание нового подменю
#                     submenu_obj = schemas.SubmenuCreate(title=submenu_data["title"], description=submenu_data["description"])
#                     new_submenu = await submenu_repo.create_submenu(new_menu.id, submenu_obj)
#                 else:
#                     # Обновление существующего подменю
#                     submenu_obj = schemas.SubmenuCreate(title=submenu_data["title"], description=submenu_data["description"])
#                     new_submenu = await submenu_repo.update_submenu(new_menu.id, existing_submenu.id, submenu_obj)
#
#                 # Обработка блюд
#                 dishes = submenu_data.get("dishes", [])
#                 for dish_data in dishes:
#                     dish_title = dish_data["title"]
#                     if existing_submenu:
#                         existing_dish = next((dish for dish in existing_submenu.dishes if dish.title == dish_title),
#                                              None)
#                     else:
#                         existing_dish = None
#
#                     if not existing_dish:
#                         # Создание нового блюда
#                         dish_obj = schemas.DishCreate(title=dish_data["title"], description=dish_data["description"], price=dish_data["price"])
#                         await dish_repo.create_dish(new_submenu.id, dish_obj)
#                     else:
#                         # Обновление существующего блюда
#                         dish_obj = schemas.DishCreate(title=dish_data["title"], description=dish_data["description"], price=dish_data["price"])
#                         await dish_repo.update_dish(new_submenu.id, existing_dish.id, dish_obj)
